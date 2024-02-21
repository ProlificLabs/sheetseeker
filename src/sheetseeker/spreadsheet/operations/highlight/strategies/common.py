import re
from typing import List, Any
from openpyxl.worksheet.worksheet import Worksheet

def prep_data_from_worksheet(sheet: Worksheet) -> List[List[Any]]:
    prepped_data = []
    curr_col_itor = 'A'
    prefix = ''
    sheet_headers = ["ROW0: COLUMN LABELS"]
    # insert the header row
    for col in range(0, sheet.max_column):
        if col == 26:
            prefix = 'A'
            curr_col_itor = 'A'
        if col > 26 and (col % 26 == 0):
            prefix = chr(ord(prefix)+1)
            curr_col_itor = 'A'
        sheet_headers.append(prefix+curr_col_itor)
        curr_col_itor = chr(ord(curr_col_itor)+1)

    prepped_data.append(sheet_headers)

    curr_row_itor = 1
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column,):
        curr_output_row = []
        for cell in row:
            if isinstance(cell.value, (bytes, bytearray, str)):
                curr_output_row.append(re.sub(r',', '', cell.value))
            else:
                curr_output_row.append(cell.value)

        if not all(cell is None for cell in curr_output_row):
            curr_output_row.insert(0, f"ROW{curr_row_itor}: ")
            prepped_data.append(curr_output_row)

        curr_row_itor+=1

    return prepped_data

def parse_llm_output_cell_ranges(output: str):
    try:
        match = re.search('\\[(.+)\\]', output)
        if match:
            output = match.group(0)[1:-1] # strip out brackets
            return [convert_row_to_range(h.strip()) for h in output.split(',')]

    except:
        raise AssertionError(f"Error, was unable to parse LLM output -> {output}")

def convert_row_to_range(input):
    match = re.search(r'^ROW(\d+)', input)
    if match:
        return f"{match.group(1)}:{match.group(1)}"
    else:
        return input
