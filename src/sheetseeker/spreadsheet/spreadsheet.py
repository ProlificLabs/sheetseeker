from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill

HIGHLIGHT_COLOR = "fffc8d"

class Spreadsheet:
    ''' This class wraps our concept of a spreadsheet. Internally it wraps an openpyxl workbook
    '''
    def __init__(self, sheet_filepath: str):
        try:
            self.sheet_filepath = sheet_filepath
            self.workbook = load_workbook(filename=sheet_filepath)
        except:
            raise ValueError("There was a problem loading the given spreadsheet file")

    def sheet_chunks(self, sheet: Worksheet):
        '''
        Split the given sheet into chunks while maintaining absolute row and
        column information from the original sheet.

        Considering this out of scope

        '''
        return sheet

    def highlight(self, worksheet_num, cells_to_highlight: List[str]):
        ''' given a worksheet number and a list of cells to highlight, highlight
        the cells (set the fgColor to yellow). An item in the list of cells is
        of the form A1, B3, to denote individual cells, or A1:A5 to denote a
        range of cells. Diagonally spanning ranges are also supported
        '''
        if not cells_to_highlight:
            return
        for cell_or_range in cells_to_highlight:
            cell_or_cells = self.sheets[worksheet_num][cell_or_range]
            if not isinstance(cell_or_cells, tuple):
                cell_or_cells = (cell_or_cells,)

            for cell in flatten(cell_or_cells):
                if not cell.value==None: # don't highlight empty cells ?
                    cell.fill = PatternFill("solid", fgColor=HIGHLIGHT_COLOR)


    def save(self, new_filename):
        ''' Save the spreadsheet to a new file using the given filename
        '''
        return self.workbook.save(new_filename)

    @property
    def original_filepath(self):
        return self.sheet_filepath;

    @property
    def sheets(self):
        return self.workbook._sheets

    @property
    def workbook(self):
        return self._workbook

    @workbook.setter
    def workbook(self, new_workbook):
        self._workbook = new_workbook

def flatten(tup):
    for item in tup:
        if isinstance(item, (tuple, list)):
            yield from flatten(item)
        else:
            yield item
