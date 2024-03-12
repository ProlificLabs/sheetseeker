import os
import math
from typing import Dict, List
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, ExcelFile

from utils import get_relevant_rows


class SheetAnalyzer:
    """Class to analyze and highlight rows in an Excel spreadsheet based on specific queries."""

    def __init__(self, file_path: str):
        """
        Initializes the SheetAnalyzer with the given file path.

        Args:
            file_path (str): The path to the Excel file to be analyzed.
        """
        self.file_path: str = file_path
        self.fill: PatternFill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.sheets: Dict[str, DataFrame] = {}
        self.queries: List[str] = [
            "Revenue",
            "Cost of Goods Sold (COGS)",
            "Operating Expenses",
            "Other Income/Expenses",
            "Interest Expense",
            "Depreciation and Amortization",
            "Stock-Based Compensation",
            "Capital Expenditures",
            "Gain/Loss on Assets",
        ]

    def find_and_highlight(self, query: str, rows_to_highlight: Dict[str, set]) -> None:
        """
        Finds and highlights the rows in the spreadsheet that match the query.

        Args:
            query (str): The query string to match against the rows.
            rows_to_highlight (Dict[str, set]): A dictionary with sheet names as keys and sets of row labels to highlight.
        """
        wb = load_workbook(self.file_path)
        for sheet_name, sheet_df in self.sheets.items():
            sheet: Worksheet = wb[sheet_name]

            for index, row in sheet_df.iterrows():
                if str(row.iloc[0]).lower() in rows_to_highlight[sheet_name]:
                    for col_index, value in enumerate(row):
                        if isinstance(value, (int, float)) and not math.isnan(value):
                            cell = sheet.cell(row=index + 2, column=col_index + 1)
                            cell.fill = self.fill

        os.makedirs('highlighted_data', exist_ok=True)
        formatted_query: str = query.replace(" ", "_").replace("/", "_")
        wb.save(f'highlighted_data/{formatted_query}_highlighted.xlsx')

    def process_query(self, query: str) -> None:
        """
        Processes a given query by finding relevant rows and highlighting them.

        Args:
            query (str): The query string to process.
        """
        rows_to_highlight: Dict[str, set] = {}
        for sheet_name, sheet_df in self.sheets.items():
            first_col: List[str] = sheet_df.iloc[:, 0].astype(str).tolist()
            relevant_rows: List[str] = get_relevant_rows(first_col, query)
            print(f"Relevant rows in {sheet_name}: {relevant_rows}")
            rows_to_highlight[sheet_name] = {row.lower() for row in relevant_rows}
        
        self.find_and_highlight(query, rows_to_highlight)

    def analyze_spreadsheet(self) -> None:
        """
        Analyzes the spreadsheet by processing a set of predefined queries.
        """
        xls: ExcelFile = pd.ExcelFile(self.file_path)
        sheet_names: List[str] = xls.sheet_names
        print("Sheet names in the Excel file:", sheet_names)
        self.sheets = {sheet_name: xls.parse(sheet_name) for sheet_name in sheet_names}

        for query in self.queries:
            print(f"\nProcessing query: {query}")
            self.process_query(query)

if __name__ == "__main__":
    file_path: str = os.path.join('data', 'sample_input.xlsx')
    analyzer: SheetAnalyzer = SheetAnalyzer(file_path)
    analyzer.analyze_spreadsheet()