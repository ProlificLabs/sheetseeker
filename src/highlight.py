from typing import List

from datatypes import MetricAnswer
import os
from constants import BASE_DIR
from parse_sheet import spreadsheet_file_path

import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def copy_xlsx_file(source_file, destination_file):
    # Copies the source XLSX file to a new location with a different name
    shutil.copy(source_file, destination_file)


def color_cells_in_xlsx(xlsx_file, sheet_name, cells, color):
    wb = load_workbook(xlsx_file)
    sheet = wb[sheet_name]
    fill_color = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell_id in cells:
        cell = sheet[cell_id]
        cell.fill = fill_color
    wb.save(xlsx_file)


def create_highlighted_excel_file(metric_answers: List[MetricAnswer]):
    """
    Given the output MetricAnswer objects, we create sheets for each target metric, and highlight the right cells

    :param metric_answers:
    :return:
    """
    if not metric_answers:
        print("Error: LLM unable to answer query, no Excel file created")
        return

    data_folder_path = os.path.join(BASE_DIR, "data")
    new_file_paths = []

    for metric_answer in metric_answers:
        # Create a new Excel file (per metric) with the relevant cells highlighted
        metric = metric_answer.metric

        # stringifying the metric to prevent any issues with file names
        metric_stringified = metric.replace("/", "_").replace("\\", "_")
        new_file_path = os.path.join(data_folder_path, f"highlighted_data_{metric_stringified}.xlsx")
        copy_xlsx_file(spreadsheet_file_path, new_file_path)

        # highlight data in new file path
        sheet = metric_answer.relevant_sheet.split(".csv")[0]
        color_cells_in_xlsx(
            new_file_path,
            sheet,
            metric_answer.relevant_cells,
            "FFFF00"
        )

        new_file_paths.append(new_file_path)

    # print all the new files created for debugging purposes
    return new_file_paths
