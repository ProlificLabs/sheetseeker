from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles.colors import Color
from langchain.chains.base import Chain
from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate
from langchain.output_parsers.openai_tools import JsonOutputKeyToolsParser


@dataclass
class FinancialDataHighlighter:
    """An LLM powered analyst that can process multi-sheet XLSX workbooks and highlight any cells that contain information relevant to a user's query."""
    system_prompt: str = """You are a financial advisor helping users find information in tables. You will be given a financial statement table and a particular piece of information to find in that table. You can use the highlight_cells tool to help the user find the relevant information in the table.
    ## General Guidelines
    * The table might be formatted in different ways, but there will likely be labels for kinds of data on a header row and column, and corresponding data points further on that row or column.
    * Your task is to highlight all cells that contain the kind of data that the user is looking for, using the provided highlight_cells tool.
    * The data with the exact title user asked for might not be present. In those cases, highlight any cell that might be relevant or that might help derive the target data.
    * Assume that this is the only document the user has, and do your best to help them get as close as they can to the information they need.
    * If multiple cells contain the kind of data the user is asking for (for example the revenue data for different years or quarters), highlight all of them.
    * Be sure to highlight **ALL** the individual cells that contain the data points, across any time scale (unless a specific time range is specified by the user).
    * You cannot chat with the user directly, and you cannot ask them whether they want you to do something different. When in doubt, highlight the most relevant data you can find."""
    model_name: str = 'gpt-4' # OpenAI model name to be used, also supports gpt-3.5-turbo
    chain: Chain = field(init=False) # Full Langchain chain that handles LLM calls

    @staticmethod
    def highlight_cells(cells: list[list[int, int]]):
        """Highlight the given cells as containing data of interest.

        Args:
            cells: array of arrays of integers, where each element corresponds to a cell: the first element of each inner array corresponds to the row and the second element corresponds to the column of the cell. Formatted like this example: `[[5,1], [5,2], [7,1], [7,2]]` This argument **CANNOT** be an empty array.
        """
        pass

    def __post_init__(self):
        self.chain = self._make_chain()

    def _make_chain(self) -> Chain:
        """
        Creates the LangChain Chain that handles formatting the prompts, calling the OpenAI model with the input parameters and parsing the tool call responses.
        """
        llm = ChatOpenAI(model_name='gpt-4').bind_tools([self.highlight_cells])
        human_template = "The user is looking for {data_title} in this table:\n{table_string}"
        prompt = ChatPromptTemplate.from_messages([
            ('system', self.system_prompt),
            ('human', human_template),
        ])
        parser = JsonOutputKeyToolsParser(key_name="highlight_cells")
        return prompt | llm | parser

    @staticmethod
    def _ws_to_prompt_string(ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
        """
        Formats the given worksheet as an LLM prompt-friendly string. To do so:
            - It iterates over the maximum range of non-empty rows and columns
            - Uses markdown table formatting to represent column and row delimiters
            - Adds header rows with integer row and column labels so the model can reliably locate cells
        This method migt cause OOMs or be inefficient for extremely sparse or pathological inputs.
        """
        res_str = '|' + '|'.join(str(i) for i in range(ws.max_column+1)) + '|\n' 
        res_str += '|' + '|'.join('---' for i in range(ws.max_column+1)) + '|\n'
        for row in range(1, ws.max_row+1):
            res_str += f'|{row}|' + '|'.join(str(ws.cell(row=row, column=col).value).replace('\n', '') if ws.cell(row=row, column=col).value else '' for col in range(1, ws.max_column+1)) + '|\n'
        res_str.split('\n')
        return res_str

    def _get_worksheet_responses(self, query:str, wb:openpyxl.Workbook) -> dict[str, list[list[list[str]]]]:
        """
        Invokes the toolchain on each of the worksheets of the given workbook for the given query, returning the raw chain output for each worksheet.
        Args:
            query: the user query for the data they're looking for
            wb: An openpyxl notebook containing data of interest
        Returns:
            A dictionary from worksheet names to parsed tool calls potentially containing the cells that should be highlighted in that worksheet.
        """
        return { sheet_name: self.chain.invoke({
            'data_title': query, 'table_string':self._ws_to_prompt_string(wb[sheet_name])
        }) for sheet_name in wb.sheetnames }

    @staticmethod
    def _highlight_cell(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int) -> None:
        """
        Given a worksheet and cell coordinates, apply the appropriate styling to highlight the cell.
        Args:
            ws: An openpyxl Worksheet
            row: Integer index of the row of the cell being highlighted
            column: Integer index of the column of the cell being highlighted
        Returns:
            None, the changes are applied directly to the ws object.
        """
        print(f"Highlighted {row},{column} in {ws.title}")
        cell = ws.cell(row=row, column=column)
        cell.style = 'Accent1' # TODO: This is a default accent style from Openpyxl, can be configured for best end user experience

    @staticmethod
    def _extract_cells_to_highlight_from_tool_call(tool_call: list[dict[str, list[list[str]]]]) -> list[tuple[int, int]]:
        """
        Parses the nested tool call object to exract tuples of integers representing the cells that should be highlighted. Handles empty tool calls.
        Args:
            tool_call: A list of dictionaries from strings to lists of lists of strings. 
                The outer list corresponds to individual tool calls for the invocation, but we just assume one tool call per invocation.
                The dictionary is a dictionary of arguments to each tool call, we extract the `cells` key from this one.
                The inner nested list is really a list of tuples for the coordinates of all the cells, so we convert it to a list of tuples of ints.
        Returns:
            list[tuple[int, int]]: All the cells that should be highlighted given this tool call.
        """
        cells_to_highlight = tool_call[0].get('cells', []) if tool_call else []
        return [(int(cell[0]), int(cell[1])) for cell in cells_to_highlight]

    def _get_worksheet_cells(self, query:str, wb:openpyxl.Workbook) -> dict[str, list[tuple[int, int]]]:
        """
        Extracts the fully parsed list of cells that need to be highlighted for each worksheet of the given workbook for the given user query.
        Args:
            query: User query specifying data of interest
            wb: Workbook potentially containing the data
        Returns:
            dict[str, list[tuple[int, int]]]: A mapping from worksheet names to lists of all the cells that should be highlighted in those worksheets.
        """
        ws_tool_calls = self._get_worksheet_responses(query, wb)
        return { 
            sheet_name: self._extract_cells_to_highlight_from_tool_call(tool_call) \
                  for sheet_name, tool_call in ws_tool_calls.items()
            }

    def process_query(self, query:str, wb:openpyxl.Workbook):
        """
        Given a user query and an openpyxl XLSX workbook, highlights all cells in the workbook that contain the data specified in the user query.
        Args:
            query: String containing the user query that specifies the data of interest
            wb: Openpyxl XLSX workbook that potentially contains the data
        Returns:
            None, the cell highlight changes are directly propagated to the workbook object
        """
        cells_to_highlight = self._get_worksheet_cells(query, wb)
        for sheet_name, cells in cells_to_highlight.items():
            for row, col in cells:
                self._highlight_cell(ws=wb[sheet_name], row=row, column=col)
        
    def analyze_workbook(self, query:str, input_path:str, output_path:str):
        """
        Given a user query and the string path to a workbook file, analyze the workbook, find cells that contain the data the user is looking for and highlight them in the workbook, saving the highlighted workbook back to disk.
        Args:
            query: String containing the user query that specifies the data of interest
            input_path: Path to an XLSX workbook that potentially contains the data
            output_path: Path to save the highlighted workbook to.
        Returns:
            None, the highlighted workbook is saved to the output path.
        """
        print(f"Analyzing {input_path}")
        wb = openpyxl.load_workbook(input_path)
        self.process_query(query, wb)
        print(f"Saving to {output_path}")
        wb.save(output_path)
        wb.close()

